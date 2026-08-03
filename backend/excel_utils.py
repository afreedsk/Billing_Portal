import re
from datetime import datetime, date
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Canonical column order used for export, and the field each maps to.
EXPORT_COLUMNS = [
    ("Date", "entry_date"),
    ("Name of the Patient", "patient_name"),
    ("Name of the Test", "test_name"),
    ("Total Amount Paid", "total_amount_paid"),
    ("Name of the Employee", "employee_name"),
    ("Cash", "cash"),
    ("Online", "online"),
    ("Paid to Other Labs", "paid_to_other_labs"),
    ("RMP", "rmp"),
    ("Salaries/Expense", "salaries_expense"),
    ("Details of Expenses", "expense_details"),
    ("Referral By", "referral_by"),
    ("Amount of Referral", "referral_amount"),
    ("Sales", "sales"),
]

CURRENCY_FIELDS = {
    "total_amount_paid", "cash", "online", "paid_to_other_labs",
    "rmp", "salaries_expense", "referral_amount", "sales",
}

HEADER_ALIASES = {
    "date": "entry_date",
    "name of the patiant": "patient_name",
    "name of the patient": "patient_name",
    "patient name": "patient_name",
    "name of the test": "test_name",
    "test name": "test_name",
    "total amount paid": "total_amount_paid",
    "name of the empolye": "employee_name",
    "name of the employee": "employee_name",
    "employee name": "employee_name",
    "cash": "cash",
    "online": "online",
    "paid to other labs": "paid_to_other_labs",
    "rmp": "rmp",
    "salaries/expense": "salaries_expense",
    "salaries expense": "salaries_expense",
    "detalis of expenses": "expense_details",
    "details of expenses": "expense_details",
    "referal by the who": "referral_by",
    "referral by the who": "referral_by",
    "referral by": "referral_by",
    "referral by who": "referral_by",
    "amount of the referal": "referral_amount",
    "amount of the referral": "referral_amount",
    "amount of referral": "referral_amount",
    "referral amount": "referral_amount",
    "sales": "sales",
}


def _normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _match_field(normalized_header):
    if normalized_header in HEADER_ALIASES:
        return HEADER_ALIASES[normalized_header]
    h = normalized_header
    if h == "date":
        return "entry_date"
    if "patiant" in h or "patient" in h:
        return "patient_name"
    if "test" in h:
        return "test_name"
    if "total" in h and "paid" in h:
        return "total_amount_paid"
    if "empolye" in h or "employee" in h:
        return "employee_name"
    if h == "cash":
        return "cash"
    if h == "online":
        return "online"
    if "other lab" in h:
        return "paid_to_other_labs"
    if h == "rmp":
        return "rmp"
    if "detail" in h and "expense" in h:
        return "expense_details"
    if "salar" in h or ("expense" in h and "detail" not in h):
        return "salaries_expense"
    if "referal" in h or "referral" in h:
        return "referral_amount" if "amount" in h else "referral_by"
    if h == "sales":
        return "sales"
    return None


def parse_excel_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_amount(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^\d.\-]", "", str(value))
    if not cleaned or cleaned in ("-", "."):
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_lab_entries_workbook(file_stream):
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_field_map = {}
    for idx, header in enumerate(header_row):
        field = _match_field(_normalize_header(header))
        if field:
            col_field_map[idx] = field

    if "entry_date" not in col_field_map.values() or "patient_name" not in col_field_map.values():
        return [], ["Could not find 'Date' and 'Name of the Patient' columns in the uploaded sheet."]

    rows = []
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {}
        for idx, field in col_field_map.items():
            value = row[idx] if idx < len(row) else None
            if field == "entry_date":
                record[field] = parse_excel_date(value)
            elif field in CURRENCY_FIELDS:
                record[field] = parse_amount(value)
            else:
                record[field] = (str(value).strip() if value is not None else None) or None
        if not record.get("patient_name") and not record.get("entry_date"):
            continue

        if not record.get("entry_date"):
            errors.append(f"Row {row_num}: missing or unreadable Date — skipped.")
            continue
        if not record.get("patient_name"):
            errors.append(f"Row {row_num}: missing Name of the Patient — skipped.")
            continue
        if not record.get("test_name"):
            record["test_name"] = "—"

        for field in CURRENCY_FIELDS:
            record.setdefault(field, 0.0)

        rows.append(record)

    return rows, errors


def build_lab_entries_workbook(entries):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lab Data Entries"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5DD4", end_color="2F5DD4", fill_type="solid")
    body_font = Font(name="Arial", size=10)
    currency_format = '"₹" #,##0.00'
    date_format = "dd-mm-yyyy"

    for col_idx, (label, _field) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, entry in enumerate(entries, start=2):
        data = entry.to_dict()
        for col_idx, (_label, field) in enumerate(EXPORT_COLUMNS, start=1):
            value = data.get(field)
            cell = ws.cell(row=row_idx, column=col_idx)
            if field == "entry_date":
                cell.value = entry.entry_date
                cell.number_format = date_format
            elif field in CURRENCY_FIELDS:
                cell.value = value or 0
                cell.number_format = currency_format
            else:
                cell.value = value or ""
            cell.font = body_font

    widths = [12, 22, 26, 16, 24, 12, 12, 16, 10, 16, 26, 20, 16, 12]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
FINANCE_HEADER_ALIASES = {
    "date": "entry_date",
    "type": "entry_type",
    "entry type": "entry_type",
    "category": "category",
    "generated by": "generated_by",
    "generated by (employee name)": "generated_by",
    "employee name": "generated_by",
    "revenue type": "revenue_type",
    "patient name": "patient_name",
    "name of the patient": "patient_name",
    "patient place": "patient_place",
    "place": "patient_place",
    "location": "patient_place",
    "client name": "client_name",
    "name of the client": "client_name",
    "client": "client_name",
    "amount": "amount",
    "amount (inr)": "amount",
    "remarks": "remarks",
    "notes": "remarks",
}

FINANCE_FIELD_LABELS = {
    "entry_date": "Date",
    "entry_type": "Type",
    "category": "Category",
    "generated_by": "Generated By",
    "revenue_type": "Revenue Type",
    "patient_name": "Patient Name",
    "patient_place": "Patient Place",
    "client_name": "Client Name",
    "amount": "Amount",
    "remarks": "Remarks",
}


def _finance_columns_for_config(config):
    columns = ["entry_date", "entry_type", "category"]
    if config.get("show_client_name"):
        columns.append("client_name")
    if config.get("show_generated_by"):
        columns.append("generated_by")
    if config.get("show_revenue_type"):
        columns.append("revenue_type")
    if config.get("show_patient_fields"):
        columns.append("patient_name")
        columns.append("patient_place")
    columns.append("amount")
    columns.append("remarks")
    return columns


def parse_finance_entries_workbook(file_stream, config, entry_types, allowed_categories_by_type):
    wb = openpyxl.load_workbook(file_stream, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_field_map = {}
    for idx, header in enumerate(header_row):
        normalized = _normalize_header(header)
        field = FINANCE_HEADER_ALIASES.get(normalized)
        if field:
            col_field_map[idx] = field

    if "entry_date" not in col_field_map.values() or "category" not in col_field_map.values():
        return [], ["Could not find 'Date' and 'Category' columns in the uploaded sheet."]

    rows = []
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {}
        for idx, field in col_field_map.items():
            value = row[idx] if idx < len(row) else None
            if field == "entry_date":
                record[field] = parse_excel_date(value)
            elif field == "amount":
                record[field] = parse_amount(value)
            else:
                record[field] = (str(value).strip() if value is not None else None) or None

        # Skip fully blank rows.
        if not record.get("entry_date") and not record.get("category"):
            continue

        if not record.get("entry_date"):
            errors.append(f"Row {row_num}: missing or unreadable Date — skipped.")
            continue
        entry_type = record.get("entry_type")
        if entry_type not in entry_types:
            matched = None
            for t in entry_types:
                if entry_type and t.lower() == str(entry_type).strip().lower():
                    matched = t
                    break
            record["entry_type"] = matched or entry_types[0]

        allowed = allowed_categories_by_type.get(record["entry_type"], [])
        if record.get("category") not in allowed:
            errors.append(
                f"Row {row_num}: category '{record.get('category')}' is not valid for "
                f"{record['entry_type']} — skipped."
            )
            continue

        if not record.get("amount") or record["amount"] <= 0:
            errors.append(f"Row {row_num}: missing or invalid Amount — skipped.")
            continue

        rows.append(record)

    return rows, errors


def build_finance_entries_workbook(entries, config, department):
    columns = _finance_columns_for_config(config)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{department} Finance Entries"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5DD4", end_color="2F5DD4", fill_type="solid")
    body_font = Font(name="Arial", size=10)
    currency_format = '"₹" #,##0.00'
    date_format = "dd-mm-yyyy"

    for col_idx, field in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=FINANCE_FIELD_LABELS[field])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, entry in enumerate(entries, start=2):
        data = entry.to_dict()
        for col_idx, field in enumerate(columns, start=1):
            value = data.get(field)
            cell = ws.cell(row=row_idx, column=col_idx)
            if field == "entry_date":
                cell.value = entry.entry_date
                cell.number_format = date_format
            elif field == "amount":
                cell.value = value or 0
                cell.number_format = currency_format
            else:
                cell.value = value or ""
            cell.font = body_font

    widths = {
        "entry_date": 12, "entry_type": 12, "category": 20, "generated_by": 22,
        "revenue_type": 16, "patient_name": 22, "patient_place": 20,
        "client_name": 22, "amount": 14, "remarks": 30,
    }
    for col_idx, field in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(field, 16)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer